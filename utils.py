# utils.py
import shutil
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet

def detect_header_row(ws, header_candidates):
    """
    Try to find a header row in worksheet ws that contains most header_candidates.
    Returns row index (1-based). If not found, returns 1.
    """
    max_row = ws.max_row
    for r in range(1, max_row + 1):
        row_vals = [str(ws.cell(row=r, column=c).value).strip() if ws.cell(row=r, column=c).value is not None else "" for c in range(1, ws.max_column + 1)]
        found = sum(1 for h in header_candidates if any(h.lower() in str(v).lower() for v in row_vals))
        if found >= max(1, len(header_candidates)//2):  # heuristic
            return r
    return 1

def find_first_empty_row_after(ws, start_row):
    """
    Find the first row after start_row where column A is empty (or all empty).
    """
    r = start_row + 1
    while True:
        if r > ws.max_row + 1000:
            return r
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if all(v is None or str(v).strip() == "" for v in row_vals):
            return r
        r += 1

def export_to_template_excel(df: pd.DataFrame, template_path: str, out_path: str, sheet_name: str = "FEBI", mapping: dict = None):
    """
    Copy template, open sheet_name, detect header, and write df rows under header.
    mapping: dict mapping template column name -> df column name (if None, use df columns in order)
    """
    template = Path(template_path)
    out = Path(out_path)
    shutil.copy(template, out)

    wb = load_workbook(out)
    if sheet_name not in wb.sheetnames:
        ws = wb.active
    else:
        ws = wb[sheet_name]

    # determine header row by detecting some expected labels
    header_candidates = list(mapping.keys()) if mapping else list(df.columns)
    header_row = detect_header_row(ws, header_candidates)

    # prepare write columns: get template headers from header_row cells (non-empty)
    template_headers = []
    header_cols = []
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=c).value
        if val is not None and str(val).strip() != "":
            template_headers.append(str(val).strip())
            header_cols.append(c)
    if not template_headers:
        # fallback: use df columns and start at row 1
        header_row = 1
        template_headers = list(df.columns)
        header_cols = list(range(1, len(template_headers)+1))
        # write headers
        for col_idx, h in enumerate(template_headers, start=1):
            ws.cell(row=header_row, column=col_idx, value=h)

    # compute first empty row after header_row
    start_write_row = find_first_empty_row_after(ws, header_row)

    # Build rows to write: for each dataframe row, map columns according to mapping
    for i, (_, row) in enumerate(df.iterrows()):
        write_row = start_write_row + i
        for j, th in enumerate(template_headers):
            col_idx = header_cols[j]
            if mapping and th in mapping:
                src_col = mapping[th]
                val = row.get(src_col, "")
            else:
                # try direct match by header name or use df column by position
                if th in df.columns:
                    val = row.get(th, "")
                else:
                    # fallback: take nth column from df
                    pos = j
                    if pos < len(df.columns):
                        val = row.iloc[pos]
                    else:
                        val = ""
            ws.cell(row=write_row, column=col_idx, value=val)

    wb.save(out)
    return str(out.resolve())

def export_to_pdf(df: pd.DataFrame, out_path: str, title: str = "Laporan Anggaran"):
    """
    Create a simple table PDF from df using reportlab.
    """
    doc = SimpleDocTemplate(out_path, pagesize=landscape(A4), leftMargin=20, rightMargin=20, topMargin=20, bottomMargin=20)
    styles = getSampleStyleSheet()
    elems = []

    elems.append(Paragraph(f"<b>{title}</b>", styles["Title"]))
    elems.append(Spacer(1, 12))

    # build table data (with header)
    data = [list(df.columns)]
    # convert all values to string (and limit length)
    for _, row in df.iterrows():
        r = []
        for v in row:
            if pd.isna(v):
                r.append("")
            else:
                s = str(v)
                if len(s) > 200:
                    s = s[:197] + "..."
                r.append(s)
        data.append(r)

    # style table
    table = Table(data, repeatRows=1)
    style = TableStyle([
        ('GRID', (0,0), (-1,-1), 0.25, colors.black),
        ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
    ])
    table.setStyle(style)
    elems.append(table)
    doc.build(elems)
    return str(Path(out_path).resolve())
